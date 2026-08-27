from __future__ import annotations

import ast
import copy
import csv
import dataclasses
import importlib
import importlib.util
import inspect
import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import numpy as np
import torch
from openpyxl import load_workbook

import recognition.realtime.knee42_ivcam as knee42_ivcam
import recognition.realtime.realtime_infer_daily30_sentence as legacy_runtime
from recognition.evaluation.team_test_report import export_team_test_reports
from recognition.evaluation.team_test_session import TeamTestSession
from recognition.inference.daily30_sentence_realtime_utils import save_prediction_logs


def load_probability_reporting(test_case: unittest.TestCase):
    module_name = "recognition.realtime.probability_reporting"
    test_case.assertIsNotNone(
        importlib.util.find_spec(module_name),
        f"missing shared probability policy module {module_name}",
    )
    return importlib.import_module(module_name)


def _team_record(*, schema_version: int, raw_probability: float) -> dict[str, object]:
    record: dict[str, object] = {
        "tester_id": "tester",
        "timestamp": "2026-08-27T00:00:00",
        "global_trial_number": 1,
        "expected_label": "A",
        "expected_text": "甲",
        "trial_number": 1,
        "predicted_label": "A",
        "predicted_text": "甲",
        "outcome": "prediction",
        "top1_correct": True,
        "top3_hit": True,
        "clip_start_sec": None,
        "clip_end_sec": None,
        "finalize_sec": None,
        "segment_duration_sec": None,
        "finalize_delay_sec": None,
        "finalize_reason": "",
        "model_version": "legacy",
    }
    if schema_version == 1:
        record.update(
            raw_confidence=raw_probability,
            calibrated_confidence=0.95,
            top3_candidates=[
                {"label": "A", "text": "甲", "confidence": raw_probability}
            ],
        )
    else:
        record.update(
            raw_probability=raw_probability,
            probability_policy={
                "kind": "uncalibrated_softmax",
                "acceptance_policy": "disabled_no_risk_coverage_evidence",
                "calibration_artifact": None,
            },
            top3_candidates=[
                {
                    "label": "A",
                    "text": "甲",
                    "raw_probability": raw_probability,
                }
            ],
        )
    return record


def _team_progress(
    *,
    schema_version: object,
    record: object,
    runtime_metadata: object,
) -> dict[str, object]:
    payload = {
        "schema_version": schema_version,
        "tester_id": "tester",
        "labels": ["A"],
        "label_display": {"A": "甲"},
        "trials_per_label": 1,
        "model_version": "legacy",
        "runtime_metadata": runtime_metadata,
        "records": [record],
    }
    if schema_version == 2 and type(schema_version) is int:
        payload["probability_policy"] = {
            "kind": "uncalibrated_softmax",
            "acceptance_policy": "disabled_no_risk_coverage_evidence",
            "calibration_artifact": None,
        }
    return payload


def _resume_team_progress(
    temp_dir: str,
    payload: object,
    *,
    runtime_metadata: dict[str, object] | None = None,
) -> TeamTestSession:
    session_dir = Path(temp_dir) / "team_tests" / "tester"
    session_dir.mkdir(parents=True)
    (session_dir / "progress.json").write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )
    return TeamTestSession(
        output_dir=Path(temp_dir),
        tester_id="tester",
        labels=["A"],
        label_display={"A": "甲"},
        trials_per_label=1,
        model_version="legacy",
        runtime_metadata=runtime_metadata,
        resume=True,
    )


def _invalid_team_record_mutations():
    return [
        ("raw_probability bool", lambda record: setattr(record, "raw_probability", True)),
        (
            "raw_probability NaN",
            lambda record: setattr(record, "raw_probability", float("nan")),
        ),
        (
            "raw_probability Infinity",
            lambda record: setattr(record, "raw_probability", float("inf")),
        ),
        (
            "top3 raw_probability bool",
            lambda record: record.top3_candidates[0].__setitem__(
                "raw_probability", True
            ),
        ),
        (
            "top3 raw_probability NaN",
            lambda record: record.top3_candidates[0].__setitem__(
                "raw_probability", float("nan")
            ),
        ),
        (
            "top3 raw_probability Infinity",
            lambda record: record.top3_candidates[0].__setitem__(
                "raw_probability", float("inf")
            ),
        ),
        (
            "legacy candidate key",
            lambda record: record.top3_candidates[0].__setitem__("confidence", 0.4),
        ),
        (
            "forged probability policy",
            lambda record: setattr(record, "probability_policy", {"kind": "forged"}),
        ),
        (
            "unknown candidate key",
            lambda record: record.top3_candidates[0].__setitem__(
                "unexpected", "value"
            ),
        ),
        (
            "non-JSON candidate value",
            lambda record: record.top3_candidates[0].__setitem__("text", object()),
        ),
    ]


TEAM_TRIAL_FIELDS = [
    "tester_id",
    "timestamp",
    "global_trial_number",
    "expected_label",
    "expected_text",
    "trial_number",
    "predicted_label",
    "predicted_text",
    "outcome",
    "top1_correct",
    "top3_hit",
    "raw_probability",
    "probability_policy",
    "top3_candidates",
    "clip_start_sec",
    "clip_end_sec",
    "finalize_sec",
    "segment_duration_sec",
    "finalize_delay_sec",
    "finalize_reason",
    "model_version",
]


def _set_record_field(record: object, field: str, value: object) -> None:
    if isinstance(record, dict):
        record[field] = value
    else:
        setattr(record, field, value)


def _record_candidates(record: object) -> list[dict[str, object]]:
    if isinstance(record, dict):
        return record["top3_candidates"]  # type: ignore[return-value]
    return record.top3_candidates  # type: ignore[attr-defined,no-any-return]


def _semantic_record_mutations():
    def set_empty_outcome(
        record: object,
        *,
        outcome: str,
        raw_probability: float = 0.0,
        top1_correct: bool = False,
        top3_hit: bool = False,
        predicted_text: str | None = None,
    ) -> None:
        _set_record_field(record, "outcome", outcome)
        _set_record_field(record, "top3_candidates", [])
        _set_record_field(record, "predicted_label", "未偵測")
        _set_record_field(
            record,
            "predicted_text",
            "未偵測" if predicted_text is None else predicted_text,
        )
        _set_record_field(record, "raw_probability", raw_probability)
        _set_record_field(record, "top1_correct", top1_correct)
        _set_record_field(record, "top3_hit", top3_hit)

    mutations: list[tuple[str, object]] = [
        (
            "top1_correct exact bool",
            lambda record: _set_record_field(record, "top1_correct", 1),
        ),
        (
            "top3_hit exact bool",
            lambda record: _set_record_field(record, "top3_hit", 1),
        ),
        (
            "unknown outcome",
            lambda record: _set_record_field(record, "outcome", "invented"),
        ),
        (
            "nonempty no_detection",
            lambda record: _set_record_field(record, "outcome", "no_detection"),
        ),
        (
            "four candidates",
            lambda record: _set_record_field(
                record,
                "top3_candidates",
                _record_candidates(record)
                + [{"label": "A", "text": "甲", "raw_probability": 0.05}],
            ),
        ),
        (
            "duplicate candidate label",
            lambda record: _record_candidates(record)[1].update(
                label="A", text="甲"
            ),
        ),
        (
            "unknown candidate label",
            lambda record: _record_candidates(record)[1].update(
                label="UNKNOWN", text="UNKNOWN"
            ),
        ),
        (
            "candidate display text mismatch",
            lambda record: _record_candidates(record)[1].__setitem__(
                "text", "錯誤文字"
            ),
        ),
        (
            "candidate probabilities increase",
            lambda record: _record_candidates(record)[1].__setitem__(
                "raw_probability", 0.7
            ),
        ),
        (
            "predicted label differs from top1",
            lambda record: _set_record_field(record, "predicted_label", "B"),
        ),
        (
            "predicted text differs from top1",
            lambda record: _set_record_field(record, "predicted_text", "乙"),
        ),
        (
            "record probability differs from top1",
            lambda record: _set_record_field(record, "raw_probability", 0.59),
        ),
        (
            "derived top1 correctness mismatch",
            lambda record: _set_record_field(record, "top1_correct", False),
        ),
        (
            "derived top3 hit mismatch",
            lambda record: _set_record_field(record, "top3_hit", False),
        ),
        (
            "empty prediction candidates",
            lambda record: _set_record_field(record, "top3_candidates", []),
        ),
        (
            "empty outcome nonzero probability",
            lambda record: set_empty_outcome(
                record, outcome="no_detection", raw_probability=0.1
            ),
        ),
        (
            "empty outcome top1 true",
            lambda record: set_empty_outcome(
                record, outcome="short_segment", top1_correct=True
            ),
        ),
        (
            "empty outcome top3 true",
            lambda record: set_empty_outcome(
                record, outcome="no_detection", top3_hit=True
            ),
        ),
        (
            "empty outcome display mismatch",
            lambda record: set_empty_outcome(
                record, outcome="no_detection", predicted_text="錯誤文字"
            ),
        ),
        (
            "clip order start after end",
            lambda record: _set_record_field(record, "clip_start_sec", 2.5),
        ),
        (
            "clip order end after finalize",
            lambda record: _set_record_field(record, "clip_end_sec", 3.5),
        ),
        (
            "derived duration mismatch",
            lambda record: _set_record_field(record, "segment_duration_sec", 0.5),
        ),
        (
            "derived delay mismatch",
            lambda record: _set_record_field(record, "finalize_delay_sec", 0.5),
        ),
    ]
    for field in (
        "clip_start_sec",
        "clip_end_sec",
        "finalize_sec",
        "segment_duration_sec",
        "finalize_delay_sec",
    ):
        mutations.extend(
            [
                (
                    f"{field} bool",
                    lambda record, field=field: _set_record_field(
                        record, field, True
                    ),
                ),
                (
                    f"{field} nonfinite",
                    lambda record, field=field: _set_record_field(
                        record, field, float("inf")
                    ),
                ),
                (
                    f"{field} missing while timing present",
                    lambda record, field=field: _set_record_field(
                        record, field, None
                    ),
                ),
            ]
        )
    return mutations


class ProbabilityPolicyTests(unittest.TestCase):
    def test_policy_is_immutable_and_explicitly_disables_acceptance(self):
        probability_reporting = load_probability_reporting(self)

        policy = probability_reporting.PROBABILITY_POLICY

        self.assertTrue(dataclasses.is_dataclass(policy))
        self.assertEqual(policy.kind, "uncalibrated_softmax")
        self.assertEqual(
            policy.acceptance_policy,
            "disabled_no_risk_coverage_evidence",
        )
        self.assertIsNone(policy.calibration_artifact)
        self.assertEqual(
            dataclasses.asdict(policy),
            {
                "kind": "uncalibrated_softmax",
                "acceptance_policy": "disabled_no_risk_coverage_evidence",
                "calibration_artifact": None,
            },
        )
        with self.assertRaises(dataclasses.FrozenInstanceError):
            policy.kind = "calibrated"  # type: ignore[misc]

    def test_raw_probability_validator_rejects_invalid_values(self):
        probability_reporting = load_probability_reporting(self)

        for invalid in (-0.001, 1.001, float("nan"), float("inf"), -float("inf")):
            with self.subTest(invalid=invalid):
                with self.assertRaisesRegex(ValueError, "finite|between 0 and 1"):
                    probability_reporting.validate_raw_probability(invalid)
        with self.assertRaisesRegex(TypeError, "real number"):
            probability_reporting.validate_raw_probability(True)


class LegacyProbabilityReportingTests(unittest.TestCase):
    def test_legacy_calibration_compatibility_shim_is_exact_identity(self):
        for raw_probability in (0.0, 0.037, 0.1, 0.45, 1.0):
            with self.subTest(raw_probability=raw_probability):
                self.assertEqual(
                    legacy_runtime.calibrate_confidence(raw_probability),
                    raw_probability,
                )

    def test_legacy_calibration_shim_rejects_invalid_probabilities(self):
        for invalid in (-0.001, 1.001, float("nan"), float("inf"), -float("inf")):
            with self.subTest(invalid=invalid):
                with self.assertRaisesRegex(ValueError, "finite|between 0 and 1"):
                    legacy_runtime.calibrate_confidence(invalid)
        with self.assertRaisesRegex(TypeError, "real number"):
            legacy_runtime.calibrate_confidence(True)

    def test_legacy_decision_never_rejects_a_low_raw_probability(self):
        output = legacy_runtime.decide_prediction_output("你好", 0.001, 0.999)

        self.assertEqual(output, ("你好", 0.001, 0.001))

    def test_segment_prediction_fields_and_top3_are_raw_probabilities(self):
        prediction = legacy_runtime.decode_segment_prediction(
            np.asarray([0.45, 0.35, 0.20], dtype=np.float64),
            ["A", "B", "C"],
            {"A": "甲", "B": "乙", "C": "丙"},
            0.99,
        )

        field_names = {field.name for field in dataclasses.fields(prediction)}
        self.assertEqual(prediction.display_label, "甲")
        self.assertEqual(prediction.raw_probability, 0.45)
        self.assertEqual(
            prediction.top3_label_probabilities,
            [("A", 0.45), ("B", 0.35), ("C", 0.20)],
        )
        self.assertEqual(
            prediction.top3_display_probabilities,
            [("甲", 0.45), ("乙", 0.35), ("丙", 0.20)],
        )
        self.assertIn("raw_probability", field_names)
        self.assertNotIn("raw_confidence", field_names)
        self.assertNotIn("calibrated_confidence", field_names)
        self.assertNotIn("accepted", field_names)
        self.assertFalse(hasattr(prediction, "accepted"))

    def test_legacy_overlay_labels_and_formats_the_exact_raw_probability(self):
        drawn_text: list[str] = []

        def capture_text(image, text, *_args, **_kwargs):
            drawn_text.append(str(text))
            return image

        with mock.patch.object(legacy_runtime, "draw_text", side_effect=capture_text):
            legacy_runtime.draw_overlay(
                np.zeros((220, 800, 3), dtype=np.uint8),
                "甲",
                [("甲", 0.037), ("乙", 0.1)],
                [],
                30.0,
                "手動切段",
                "等待",
                "IDLE_BLANK",
                "manual",
            )

        candidate_line = next(text for text in drawn_text if "候選結果" in text)
        self.assertIn("raw probability", candidate_line.lower())
        self.assertIn("3.7%", candidate_line)
        self.assertIn("10.0%", candidate_line)
        self.assertNotIn("15.0%", candidate_line)

    def test_legacy_nonnegative_threshold_override_fails_at_resolution(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            missing_config = Path(temp_dir) / "missing.json"
            for override in ("0", "0.5", "1"):
                with self.subTest(override=override):
                    parsed = legacy_runtime.parse_args(
                        [
                            "--app-config",
                            str(missing_config),
                            "--min-conf-override",
                            override,
                        ]
                    )
                    with self.assertRaisesRegex(
                        ValueError,
                        "acceptance threshold is unavailable.*calibration/risk-coverage evidence",
                    ):
                        legacy_runtime.resolve_runtime_args(parsed)

    def test_legacy_default_neutralizes_the_deprecated_threshold_option(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            args = legacy_runtime.resolve_runtime_args(
                legacy_runtime.parse_args(
                    ["--app-config", str(Path(temp_dir) / "missing.json")]
                )
            )

        self.assertIsNone(args.min_conf_override)

    def test_legacy_executable_contains_no_probability_inflation_formula(self):
        source = inspect.getsource(legacy_runtime)

        self.assertNotRegex(source, r"\*\*\s*0\.7(?:0)?")
        self.assertNotRegex(source, r"np\.clip\([^\n]*0\.95")
        self.assertNotIn("calibrated confidence", source.lower())


class LegacyTeamRecordProbabilityTests(unittest.TestCase):
    def test_new_team_record_schema_uses_raw_probability_and_disabled_policy(self):
        parameters = inspect.signature(TeamTestSession.stage_prediction).parameters

        self.assertIn("raw_probability", parameters)
        self.assertNotIn("raw_confidence", parameters)
        self.assertNotIn("calibrated_confidence", parameters)

        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
            )
            record = session.stage_prediction(
                predicted_label="A",
                raw_probability=0.4,
                top3_candidates=[("A", 0.4)],
            )

        payload = dataclasses.asdict(record)
        self.assertEqual(payload["raw_probability"], 0.4)
        self.assertNotIn("raw_confidence", payload)
        self.assertNotIn("calibrated_confidence", payload)
        self.assertEqual(payload["top3_candidates"][0]["raw_probability"], 0.4)
        self.assertNotIn("confidence", payload["top3_candidates"][0])
        self.assertEqual(
            payload["probability_policy"]["acceptance_policy"],
            "disabled_no_risk_coverage_evidence",
        )

    def test_schema_one_team_progress_is_converted_to_raw_probability_on_load(self):
        legacy_record = {
            "tester_id": "tester",
            "timestamp": "2026-08-27T00:00:00",
            "global_trial_number": 1,
            "expected_label": "A",
            "expected_text": "甲",
            "trial_number": 1,
            "predicted_label": "A",
            "predicted_text": "甲",
            "outcome": "prediction",
            "top1_correct": True,
            "top3_hit": True,
            "raw_confidence": 0.4,
            "calibrated_confidence": 0.95,
            "top3_candidates": [{"label": "A", "text": "甲", "confidence": 0.4}],
            "clip_start_sec": None,
            "clip_end_sec": None,
            "finalize_sec": None,
            "segment_duration_sec": None,
            "finalize_delay_sec": None,
            "finalize_reason": "",
            "model_version": "legacy",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            session_dir = Path(temp_dir) / "team_tests" / "tester"
            session_dir.mkdir(parents=True)
            (session_dir / "progress.json").write_text(
                json.dumps(
                    {
                        "schema_version": 1,
                        "tester_id": "tester",
                        "labels": ["A"],
                        "label_display": {"A": "甲"},
                        "trials_per_label": 1,
                        "model_version": "legacy",
                        "runtime_metadata": {},
                        "records": [legacy_record],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            resumed = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
                resume=True,
            )

        self.assertTrue(hasattr(resumed.records[0], "raw_probability"))
        self.assertEqual(resumed.records[0].raw_probability, 0.4)
        self.assertEqual(resumed.records[0].top3_candidates[0]["raw_probability"], 0.4)

    def test_progress_schema_version_is_strict_and_never_treats_bool_as_int(self):
        record = _team_record(schema_version=2, raw_probability=0.4)

        for invalid_schema in (True, False, None, 0, 3, "2", 2.0):
            with self.subTest(invalid_schema=invalid_schema), tempfile.TemporaryDirectory() as temp_dir:
                payload = _team_progress(
                    schema_version=invalid_schema,
                    record=record,
                    runtime_metadata={},
                )
                with self.assertRaisesRegex(ValueError, "schema_version"):
                    _resume_team_progress(temp_dir, payload)

    def test_schema_two_rejects_legacy_probability_keys_in_records_and_top3(self):
        mutations = {
            "confidence": lambda record: record.__setitem__("confidence", 0.4),
            "raw_confidence": lambda record: record.__setitem__("raw_confidence", 0.4),
            "calibrated_confidence": lambda record: record.__setitem__(
                "calibrated_confidence", 0.95
            ),
            "top3 confidence": lambda record: record["top3_candidates"][0].__setitem__(
                "confidence", 0.4
            ),
        }

        for name, mutate in mutations.items():
            with self.subTest(name=name), tempfile.TemporaryDirectory() as temp_dir:
                record = _team_record(schema_version=2, raw_probability=0.4)
                mutate(record)
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                with self.assertRaisesRegex(ValueError, "legacy probability key"):
                    _resume_team_progress(temp_dir, payload)

    def test_progress_rejects_non_objects_bool_and_nonfinite_probabilities(self):
        cases: list[tuple[str, object]] = []
        for invalid_probability in (True, float("nan"), float("inf"), -float("inf")):
            record = _team_record(schema_version=2, raw_probability=0.4)
            record["raw_probability"] = invalid_probability
            cases.append((f"raw probability {invalid_probability!r}", record))
        cases.extend(
            [
                ("record non-object", "not-an-object"),
                (
                    "top3 item non-object",
                    {
                        **_team_record(schema_version=2, raw_probability=0.4),
                        "top3_candidates": ["not-an-object"],
                    },
                ),
            ]
        )

        for name, record in cases:
            with self.subTest(name=name), tempfile.TemporaryDirectory() as temp_dir:
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                with self.assertRaisesRegex(ValueError, "record|top-3|raw probability"):
                    _resume_team_progress(temp_dir, payload)

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "JSON object"):
                _resume_team_progress(temp_dir, ["not", "an", "object"])

    def test_real_schema_one_runtime_metadata_migrates_only_probability_fields(self):
        legacy_metadata = {
            "model_sha256": "abc123",
            "auto_trigger": {"end_hold_sec": 0.6},
            "temporal_model_loaded": False,
            "source": "0",
            "backend": "dshow",
            "confidence_threshold": 0.5,
        }
        current_metadata = {
            "model_sha256": "abc123",
            "auto_trigger": {"end_hold_sec": 0.6},
            "temporal_model_loaded": False,
            "source": "0",
            "backend": "dshow",
        }
        payload = _team_progress(
            schema_version=1,
            record=_team_record(schema_version=1, raw_probability=0.4),
            runtime_metadata=legacy_metadata,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            resumed = _resume_team_progress(
                temp_dir,
                payload,
                runtime_metadata=current_metadata,
            )

        self.assertEqual(resumed.records[0].raw_probability, 0.4)

        changed_metadata = copy.deepcopy(legacy_metadata)
        changed_metadata["model_sha256"] = "different"
        changed_payload = _team_progress(
            schema_version=1,
            record=_team_record(schema_version=1, raw_probability=0.4),
            runtime_metadata=changed_metadata,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "runtime_metadata"):
                _resume_team_progress(
                    temp_dir,
                    changed_payload,
                    runtime_metadata=current_metadata,
                )

    def test_schema_two_writer_rejects_legacy_runtime_metadata_at_construction(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(
                ValueError,
                "runtime_metadata.*confidence_threshold|confidence_threshold.*runtime_metadata",
            ):
                TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=1,
                    model_version="legacy",
                    runtime_metadata={"confidence_threshold": 0.5},
                )

    def test_progress_writer_rechecks_runtime_metadata_after_mutation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
                runtime_metadata={"source": "0"},
            )
            session.stage_prediction(
                predicted_label="A",
                raw_probability=0.4,
                top3_candidates=[("A", 0.4)],
            )
            session.runtime_metadata["confidence_threshold"] = 0.5

            with self.assertRaisesRegex(ValueError, "confidence_threshold"):
                session.confirm_pending()

            self.assertEqual(session.records, [])
            self.assertFalse(session.progress_path.exists())

    def test_report_writer_rechecks_runtime_metadata_after_mutation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
                runtime_metadata={"source": "0"},
            )
            session.confirm_prediction(
                predicted_label="A",
                raw_probability=0.4,
                top3_candidates=[("A", 0.4)],
            )
            session.runtime_metadata["confidence_threshold"] = 0.5

            with self.assertRaisesRegex(ValueError, "confidence_threshold"):
                export_team_test_reports(session)

            self.assertEqual(list(session.session_dir.glob("team_results_*.xlsx")), [])
            self.assertEqual(list(session.session_dir.glob("team_trials_*.csv")), [])
            self.assertEqual(list(session.session_dir.glob("team_session_*.json")), [])

    def test_resume_rejects_incompatible_label_display(self):
        payload = _team_progress(
            schema_version=2,
            record=_team_record(schema_version=2, raw_probability=0.4),
            runtime_metadata={},
        )
        payload["label_display"] = {"A": "不同文字"}

        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "label_display"):
                _resume_team_progress(temp_dir, payload)

    def test_resume_rejects_record_identity_and_label_config_mismatches(self):
        mutations = {
            "tester_id": lambda record: record.__setitem__("tester_id", "other"),
            "model_version": lambda record: record.__setitem__(
                "model_version", "other-model"
            ),
            "expected_label": lambda record: record.__setitem__(
                "expected_label", "UNKNOWN"
            ),
            "expected_text": lambda record: record.__setitem__(
                "expected_text", "不同文字"
            ),
        }

        for field, mutate in mutations.items():
            with self.subTest(field=field), tempfile.TemporaryDirectory() as temp_dir:
                record = _team_record(schema_version=2, raw_probability=0.4)
                mutate(record)
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                with self.assertRaisesRegex(ValueError, field):
                    _resume_team_progress(temp_dir, payload)

    def test_identity_types_and_bool_sequences_fail_reads_and_writes_atomically(self):
        mutations = {
            "global_trial_number": lambda record: record.__setitem__(
                "global_trial_number", True
            ),
            "trial_number": lambda record: record.__setitem__("trial_number", True),
            "tester_id": lambda record: record.__setitem__("tester_id", 123),
            "timestamp": lambda record: record.__setitem__("timestamp", 123),
            "expected_label": lambda record: record.__setitem__("expected_label", 123),
            "expected_text": lambda record: record.__setitem__("expected_text", 123),
            "predicted_label": lambda record: record.__setitem__("predicted_label", 123),
            "predicted_text": lambda record: record.__setitem__("predicted_text", 123),
            "outcome": lambda record: record.__setitem__("outcome", 123),
            "finalize_reason": lambda record: record.__setitem__("finalize_reason", 123),
            "model_version": lambda record: record.__setitem__("model_version", 123),
            "top3 label": lambda record: record["top3_candidates"][0].__setitem__(
                "label", 123
            ),
            "top3 text": lambda record: record["top3_candidates"][0].__setitem__(
                "text", 123
            ),
        }

        for field, mutate in mutations.items():
            with self.subTest(field=field), tempfile.TemporaryDirectory() as temp_dir:
                record = _team_record(schema_version=2, raw_probability=0.4)
                mutate(record)
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )

                with self.assertRaisesRegex(ValueError, field.replace(" ", ".*")):
                    _resume_team_progress(temp_dir, payload)

        mutations = {
            "global_trial_number": lambda record: setattr(
                record, "global_trial_number", True
            ),
            "trial_number": lambda record: setattr(record, "trial_number", True),
            "tester_id": lambda record: setattr(record, "tester_id", 123),
            "timestamp": lambda record: setattr(record, "timestamp", 123),
            "expected_label": lambda record: setattr(record, "expected_label", 123),
            "expected_text": lambda record: setattr(record, "expected_text", 123),
            "predicted_label": lambda record: setattr(record, "predicted_label", 123),
            "predicted_text": lambda record: setattr(record, "predicted_text", 123),
            "outcome": lambda record: setattr(record, "outcome", 123),
            "finalize_reason": lambda record: setattr(record, "finalize_reason", 123),
            "model_version": lambda record: setattr(record, "model_version", 123),
            "top3 label": lambda record: record.top3_candidates[0].__setitem__(
                "label", 123
            ),
            "top3 text": lambda record: record.top3_candidates[0].__setitem__(
                "text", 123
            ),
        }

        for field, mutate in mutations.items():
            with self.subTest(field=field), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=2,
                    model_version="legacy",
                )
                first = session.confirm_prediction(
                    predicted_label="A",
                    raw_probability=0.4,
                    top3_candidates=[("A", 0.4)],
                )
                progress_bytes = session.progress_path.read_bytes()
                pending = session.stage_prediction(
                    predicted_label="A",
                    raw_probability=0.3,
                    top3_candidates=[("A", 0.3)],
                )
                mutate(pending)

                with self.assertRaisesRegex(ValueError, field.replace(" ", ".*")):
                    session.confirm_pending()

                self.assertEqual(session.records, [first])
                self.assertIs(session.pending_result, pending)
                self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session.session_dir.glob("*.tmp")), [])

    def test_schema_two_rejects_runtime_probability_policy_shadow(self):
        canonical_policy = {
            "kind": "uncalibrated_softmax",
            "acceptance_policy": "disabled_no_risk_coverage_evidence",
            "calibration_artifact": None,
        }
        for shadow in (canonical_policy, {"kind": "forged"}):
            with self.subTest(source="writer", shadow=shadow), tempfile.TemporaryDirectory() as temp_dir:
                with self.assertRaisesRegex(ValueError, "runtime_metadata.*probability_policy"):
                    TeamTestSession(
                        output_dir=Path(temp_dir),
                        tester_id="tester",
                        labels=["A"],
                        label_display={"A": "甲"},
                        trials_per_label=1,
                        model_version="legacy",
                        runtime_metadata={"probability_policy": shadow},
                    )

            with self.subTest(source="reader", shadow=shadow), tempfile.TemporaryDirectory() as temp_dir:
                payload = _team_progress(
                    schema_version=2,
                    record=_team_record(schema_version=2, raw_probability=0.4),
                    runtime_metadata={"probability_policy": shadow},
                )
                with self.assertRaisesRegex(ValueError, "runtime_metadata.*probability_policy"):
                    _resume_team_progress(temp_dir, payload)

        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=2,
                model_version="legacy",
                runtime_metadata={"source": "0"},
            )
            first = session.confirm_prediction(
                predicted_label="A",
                raw_probability=0.4,
                top3_candidates=[("A", 0.4)],
            )
            progress_bytes = session.progress_path.read_bytes()
            pending = session.stage_prediction(
                predicted_label="A",
                raw_probability=0.3,
                top3_candidates=[("A", 0.3)],
            )
            session.runtime_metadata["probability_policy"] = canonical_policy

            with self.assertRaisesRegex(ValueError, "runtime_metadata.*probability_policy"):
                session.confirm_pending()

            self.assertEqual(session.records, [first])
            self.assertIs(session.pending_result, pending)
            self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
            self.assertEqual(list(session.session_dir.glob("*.tmp")), [])

        main_tree = ast.parse(inspect.getsource(legacy_runtime.main))
        team_session_calls = [
            node
            for node in ast.walk(main_tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Name)
            and node.func.id == "TeamTestSession"
        ]
        self.assertEqual(len(team_session_calls), 1)
        runtime_metadata_nodes = [
            keyword.value
            for keyword in team_session_calls[0].keywords
            if keyword.arg == "runtime_metadata"
        ]
        self.assertEqual(len(runtime_metadata_nodes), 1)
        self.assertIsInstance(runtime_metadata_nodes[0], ast.Dict)
        runtime_metadata_keys = {
            key.value
            for key in runtime_metadata_nodes[0].keys
            if isinstance(key, ast.Constant) and isinstance(key.value, str)
        }
        self.assertNotIn("probability_policy", runtime_metadata_keys)

        with tempfile.TemporaryDirectory() as temp_dir:
            callsite_session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
                runtime_metadata={
                    "model_sha256": "abc123",
                    "auto_trigger": {"end_hold_sec": 0.6},
                    "temporal_model_loaded": False,
                    "source": "0",
                    "backend": "dshow",
                },
            )
            callsite_payload = callsite_session._metadata()

        self.assertEqual(callsite_payload["probability_policy"], canonical_policy)
        self.assertNotIn(
            "probability_policy",
            callsite_payload["runtime_metadata"],
        )

    def test_progress_top_level_rejects_unknown_fields_in_both_schemas(self):
        for schema_version in (1, 2):
            with self.subTest(schema_version=schema_version), tempfile.TemporaryDirectory() as temp_dir:
                payload = _team_progress(
                    schema_version=schema_version,
                    record=_team_record(schema_version=schema_version, raw_probability=0.4),
                    runtime_metadata={},
                )
                payload["unknown_future_field"] = {"must_not_be_dropped": True}
                session_dir = Path(temp_dir) / "team_tests" / "tester"
                session_dir.mkdir(parents=True)
                progress_path = session_dir / "progress.json"
                progress_path.write_text(
                    json.dumps(payload, ensure_ascii=False),
                    encoding="utf-8",
                )
                progress_bytes = progress_path.read_bytes()

                with self.assertRaisesRegex(ValueError, "unknown.*unknown_future_field"):
                    TeamTestSession(
                        output_dir=Path(temp_dir),
                        tester_id="tester",
                        labels=["A"],
                        label_display={"A": "甲"},
                        trials_per_label=1,
                        model_version="legacy",
                        resume=True,
                    )

                self.assertEqual(progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session_dir.glob("*.tmp")), [])

    def test_schema_one_top3_rejects_noncanonical_probability_keys(self):
        for forbidden_key in (
            "raw_confidence",
            "calibrated_confidence",
            "raw_probability",
        ):
            with (
                self.subTest(forbidden_key=forbidden_key),
                tempfile.TemporaryDirectory() as temp_dir,
            ):
                record = _team_record(schema_version=1, raw_probability=0.4)
                record["top3_candidates"][0][forbidden_key] = 0.4
                payload = _team_progress(
                    schema_version=1,
                    record=record,
                    runtime_metadata={},
                )
                with self.assertRaisesRegex(
                    ValueError,
                    f"top-3.*{forbidden_key}|{forbidden_key}.*top-3",
                ):
                    _resume_team_progress(temp_dir, payload)

    def test_pending_record_mutations_fail_confirm_without_files_or_state_change(self):
        for name, mutate in _invalid_team_record_mutations():
            with self.subTest(name=name), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=1,
                    model_version="legacy",
                )
                pending = session.stage_prediction(
                    predicted_label="A",
                    raw_probability=0.4,
                    top3_candidates=[("A", 0.4)],
                )
                mutate(pending)

                with self.assertRaises(ValueError):
                    session.confirm_pending()

                self.assertEqual(session.records, [])
                self.assertIs(session.pending_result, pending)
                self.assertEqual(list(session.session_dir.glob("*")), [])

    def test_loaded_record_mutations_fail_export_without_partial_reports(self):
        for name, mutate in _invalid_team_record_mutations():
            with self.subTest(name=name), tempfile.TemporaryDirectory() as temp_dir:
                initial = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=1,
                    model_version="legacy",
                )
                initial.confirm_prediction(
                    predicted_label="A",
                    raw_probability=0.4,
                    top3_candidates=[("A", 0.4)],
                )
                progress_bytes = initial.progress_path.read_bytes()
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=1,
                    model_version="legacy",
                    resume=True,
                )
                mutate(session.records[0])

                with self.assertRaises(ValueError):
                    export_team_test_reports(session)

                self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session.session_dir.glob("team_results_*.xlsx")), [])
                self.assertEqual(list(session.session_dir.glob("team_trials_*.csv")), [])
                self.assertEqual(list(session.session_dir.glob("team_session_*.json")), [])

    def test_confirm_failures_leave_records_pending_and_progress_atomically_unchanged(self):
        for failure in ("validation", "serialization", "write", "replace"):
            with self.subTest(failure=failure), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "甲"},
                    trials_per_label=2,
                    model_version="legacy",
                )
                first = session.confirm_prediction(
                    predicted_label="A",
                    raw_probability=0.4,
                    top3_candidates=[("A", 0.4)],
                )
                progress_bytes = session.progress_path.read_bytes()
                pending = session.stage_prediction(
                    predicted_label="A",
                    raw_probability=0.3,
                    top3_candidates=[("A", 0.3)],
                )

                if failure == "validation":
                    pending.raw_probability = True
                    with self.assertRaises(ValueError):
                        session.confirm_pending()
                elif failure == "serialization":
                    pending.top3_candidates[0]["text"] = object()
                    with self.assertRaises(ValueError):
                        session.confirm_pending()
                elif failure == "write":
                    with mock.patch.object(
                        Path, "write_text", side_effect=OSError("write failed")
                    ):
                        with self.assertRaises(OSError):
                            session.confirm_pending()
                else:
                    with mock.patch.object(
                        Path, "replace", side_effect=OSError("replace failed")
                    ):
                        with self.assertRaises(OSError):
                            session.confirm_pending()

                self.assertEqual(session.records, [first])
                self.assertIs(session.pending_result, pending)
                self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session.session_dir.glob("*.tmp")), [])

    def test_resume_rejects_record_sequence_mismatches(self):
        mutations = {
            "global_trial_number": lambda record: record.__setitem__(
                "global_trial_number", 2
            ),
            "trial_number": lambda record: record.__setitem__("trial_number", 2),
            "expected_label": lambda record: record.update(
                expected_label="B", expected_text="乙"
            ),
        }

        for field, mutate in mutations.items():
            with self.subTest(field=field), tempfile.TemporaryDirectory() as temp_dir:
                record = _team_record(schema_version=2, raw_probability=0.4)
                mutate(record)
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                payload["labels"] = ["A", "B"]
                payload["label_display"] = {"A": "甲", "B": "乙"}
                session_dir = Path(temp_dir) / "team_tests" / "tester"
                session_dir.mkdir(parents=True)
                (session_dir / "progress.json").write_text(
                    json.dumps(payload, ensure_ascii=False), encoding="utf-8"
                )

                with self.assertRaisesRegex(ValueError, field):
                    TeamTestSession(
                        output_dir=Path(temp_dir),
                        tester_id="tester",
                        labels=["A", "B"],
                        label_display={"A": "甲", "B": "乙"},
                        trials_per_label=1,
                        model_version="legacy",
                        resume=True,
                    )

    def test_resume_requires_records_and_each_records_top3_candidates(self):
        for schema_version in (1, 2):
            for missing_field in ("records", "top3_candidates"):
                with (
                    self.subTest(
                        schema_version=schema_version,
                        missing_field=missing_field,
                    ),
                    tempfile.TemporaryDirectory() as temp_dir,
                ):
                    record = _team_record(
                        schema_version=schema_version,
                        raw_probability=0.4,
                    )
                    payload = _team_progress(
                        schema_version=schema_version,
                        record=record,
                        runtime_metadata={},
                    )
                    if missing_field == "records":
                        payload.pop("records")
                    else:
                        record.pop("top3_candidates")
                    session_dir = Path(temp_dir) / "team_tests" / "tester"
                    session_dir.mkdir(parents=True)
                    progress_path = session_dir / "progress.json"
                    progress_path.write_text(
                        json.dumps(payload, ensure_ascii=False),
                        encoding="utf-8",
                    )
                    progress_bytes = progress_path.read_bytes()

                    with self.assertRaisesRegex(ValueError, missing_field):
                        TeamTestSession(
                            output_dir=Path(temp_dir),
                            tester_id="tester",
                            labels=["A"],
                            label_display={"A": "甲"},
                            trials_per_label=1,
                            model_version="legacy",
                            resume=True,
                        )

                    self.assertEqual(progress_path.read_bytes(), progress_bytes)
                    self.assertEqual(list(session_dir.rglob("*.tmp")), [])

    def test_record_semantic_and_timing_mutations_fail_all_paths_atomically(self):
        labels = ["A", "B", "C"]
        label_display = {"A": "甲", "B": "乙", "C": "丙"}

        def stage_valid(session: TeamTestSession):
            return session.stage_prediction(
                predicted_label="A",
                raw_probability=0.6,
                top3_candidates=[("A", 0.6), ("B", 0.3), ("C", 0.1)],
                clip_start_sec=1.0,
                clip_end_sec=2.0,
                finalize_sec=3.0,
            )

        for name, mutate in _semantic_record_mutations():
            with self.subTest(name=name, path="resume"), tempfile.TemporaryDirectory() as temp_dir:
                record = _team_record(schema_version=2, raw_probability=0.6)
                record.update(
                    top3_candidates=[
                        {"label": "A", "text": "甲", "raw_probability": 0.6},
                        {"label": "B", "text": "乙", "raw_probability": 0.3},
                        {"label": "C", "text": "丙", "raw_probability": 0.1},
                    ],
                    clip_start_sec=1.0,
                    clip_end_sec=2.0,
                    finalize_sec=3.0,
                    segment_duration_sec=1.0,
                    finalize_delay_sec=1.0,
                )
                mutate(record)
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                payload["labels"] = labels
                payload["label_display"] = label_display
                session_dir = Path(temp_dir) / "team_tests" / "tester"
                session_dir.mkdir(parents=True)
                progress_path = session_dir / "progress.json"
                progress_path.write_text(
                    json.dumps(payload, ensure_ascii=False), encoding="utf-8"
                )
                progress_bytes = progress_path.read_bytes()

                with self.assertRaises(ValueError):
                    TeamTestSession(
                        output_dir=Path(temp_dir),
                        tester_id="tester",
                        labels=labels,
                        label_display=label_display,
                        trials_per_label=1,
                        model_version="legacy",
                        resume=True,
                    )

                self.assertEqual(progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session_dir.rglob("*.tmp")), [])

            with self.subTest(name=name, path="confirm"), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=labels,
                    label_display=label_display,
                    trials_per_label=1,
                    model_version="legacy",
                )
                pending = stage_valid(session)
                mutate(pending)

                with self.assertRaises(ValueError):
                    session.confirm_pending()

                self.assertEqual(session.records, [])
                self.assertIs(session.pending_result, pending)
                self.assertEqual(list(session.session_dir.rglob("*")), [])

            with self.subTest(name=name, path="export"), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=labels,
                    label_display=label_display,
                    trials_per_label=1,
                    model_version="legacy",
                )
                stage_valid(session)
                session.confirm_pending()
                progress_bytes = session.progress_path.read_bytes()
                mutate(session.records[0])

                with self.assertRaises(ValueError):
                    export_team_test_reports(session)

                self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session.session_dir.glob("team_results_*.xlsx")), [])
                self.assertEqual(list(session.session_dir.glob("team_trials_*.csv")), [])
                self.assertEqual(list(session.session_dir.glob("team_session_*.json")), [])
                self.assertEqual(list(session.session_dir.rglob("*.tmp")), [])

    def test_empty_outcome_sentinels_fail_closed_atomically_on_all_paths(self):
        invalid_sentinels = [
            ("no_detection", "A", "\u7532"),
            ("short_segment", "A", "\u7532"),
            ("no_detection", "\u7247\u6bb5\u904e\u77ed", "\u7247\u6bb5\u904e\u77ed"),
            ("short_segment", "\u672a\u5075\u6e2c", "\u672a\u5075\u6e2c"),
            ("no_detection", "\u932f\u8aa4 sentinel", "\u932f\u8aa4 sentinel"),
            ("short_segment", "\u932f\u8aa4 sentinel", "\u932f\u8aa4 sentinel"),
        ]
        expected_sentinels = {
            "no_detection": "\u672a\u5075\u6e2c",
            "short_segment": "\u7247\u6bb5\u904e\u77ed",
        }

        for outcome, predicted_label, predicted_text in invalid_sentinels:
            with (
                self.subTest(outcome=outcome, label=predicted_label, path="resume"),
                tempfile.TemporaryDirectory() as temp_dir,
            ):
                record = _team_record(schema_version=2, raw_probability=0.0)
                record.update(
                    outcome=outcome,
                    predicted_label=predicted_label,
                    predicted_text=predicted_text,
                    top1_correct=False,
                    top3_hit=False,
                    raw_probability=0.0,
                    top3_candidates=[],
                )
                payload = _team_progress(
                    schema_version=2,
                    record=record,
                    runtime_metadata={},
                )
                session_dir = Path(temp_dir) / "team_tests" / "tester"
                session_dir.mkdir(parents=True)
                progress_path = session_dir / "progress.json"
                progress_path.write_text(
                    json.dumps(payload, ensure_ascii=False), encoding="utf-8"
                )
                progress_bytes = progress_path.read_bytes()

                with self.assertRaisesRegex(ValueError, "sentinel"):
                    TeamTestSession(
                        output_dir=Path(temp_dir),
                        tester_id="tester",
                        labels=["A"],
                        label_display={"A": "\u7532"},
                        trials_per_label=1,
                        model_version="legacy",
                        resume=True,
                    )

                self.assertEqual(progress_path.read_bytes(), progress_bytes)
                self.assertEqual(list(session_dir.rglob("*.tmp")), [])

            with (
                self.subTest(outcome=outcome, label=predicted_label, path="confirm"),
                tempfile.TemporaryDirectory() as temp_dir,
            ):
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "\u7532"},
                    trials_per_label=1,
                    model_version="legacy",
                )
                pending = session.stage_prediction(
                    predicted_label=expected_sentinels[outcome],
                    raw_probability=0.0,
                    top3_candidates=[],
                    outcome=outcome,
                )
                pending.predicted_label = predicted_label
                pending.predicted_text = predicted_text
                pending.top1_correct = False

                with self.assertRaisesRegex(ValueError, "sentinel"):
                    session.confirm_pending()

                self.assertEqual(session.records, [])
                self.assertIs(session.pending_result, pending)
                self.assertEqual(list(session.session_dir.rglob("*")), [])

            with (
                self.subTest(outcome=outcome, label=predicted_label, path="export"),
                tempfile.TemporaryDirectory() as temp_dir,
            ):
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=["A"],
                    label_display={"A": "\u7532"},
                    trials_per_label=1,
                    model_version="legacy",
                )
                record = session.confirm_prediction(
                    predicted_label=expected_sentinels[outcome],
                    raw_probability=0.0,
                    top3_candidates=[],
                    outcome=outcome,
                )
                progress_bytes = session.progress_path.read_bytes()
                record.predicted_label = predicted_label
                record.predicted_text = predicted_text

                with self.assertRaisesRegex(ValueError, "sentinel"):
                    export_team_test_reports(session)

                self.assertEqual(session.progress_path.read_bytes(), progress_bytes)
                self.assertEqual(
                    list(session.session_dir.glob("team_results_*.xlsx")), []
                )
                self.assertEqual(
                    list(session.session_dir.glob("team_trials_*.csv")), []
                )
                self.assertEqual(
                    list(session.session_dir.glob("team_session_*.json")), []
                )
                self.assertEqual(list(session.session_dir.rglob("*.tmp")), [])

    def test_valid_prediction_top3_lengths_and_empty_outcomes_round_trip(self):
        labels = ["A", "B", "C"]
        label_display = {"A": "甲", "B": "乙", "C": "丙"}
        cases = [
            ("prediction", "A", 0.8, [("A", 0.8)]),
            ("prediction", "B", 0.6, [("B", 0.6), ("A", 0.4)]),
            (
                "prediction",
                "C",
                0.6,
                [("C", 0.6), ("A", 0.3), ("B", 0.1)],
            ),
            ("no_detection", "未偵測", 0.0, []),
            ("short_segment", "片段過短", 0.0, []),
        ]
        for outcome, predicted_label, raw_probability, candidates in cases:
            with self.subTest(outcome=outcome, candidates=len(candidates)), tempfile.TemporaryDirectory() as temp_dir:
                session = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=labels,
                    label_display=label_display,
                    trials_per_label=1,
                    model_version="legacy",
                )
                confirmed = session.confirm_prediction(
                    predicted_label=predicted_label,
                    raw_probability=raw_probability,
                    top3_candidates=candidates,
                    outcome=outcome,
                )
                resumed = TeamTestSession(
                    output_dir=Path(temp_dir),
                    tester_id="tester",
                    labels=labels,
                    label_display=label_display,
                    trials_per_label=1,
                    model_version="legacy",
                    resume=True,
                )
                reports = export_team_test_reports(resumed)

                self.assertEqual(resumed.records, [confirmed])
                self.assertEqual(confirmed.top1_correct, predicted_label == "A")
                self.assertEqual(
                    confirmed.top3_hit,
                    any(label == "A" for label, _ in candidates),
                )
                self.assertTrue(all(path.is_file() for path in dataclasses.astuple(reports)))

    def test_empty_exports_use_the_fixed_full_team_trial_header(self):
        self.assertEqual(len(TEAM_TRIAL_FIELDS), 21)
        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
            )
            paths = export_team_test_reports(session)
            with paths.trials_csv.open(encoding="utf-8-sig", newline="") as handle:
                csv_rows = list(csv.reader(handle))
            workbook = load_workbook(paths.workbook, read_only=True, data_only=True)
            xlsx_rows = list(
                workbook["逐次測試結果"].iter_rows(values_only=True)
            )
            workbook.close()

        self.assertEqual(csv_rows, [TEAM_TRIAL_FIELDS])
        self.assertEqual(xlsx_rows, [tuple(TEAM_TRIAL_FIELDS)])

    def test_legacy_log_fields_and_files_preserve_exact_probability_floats(self):
        raw_probability = 0.12345678901234568
        builder = getattr(legacy_runtime, "build_probability_log_fields", None)

        self.assertTrue(callable(builder), "missing exact probability log-field builder")
        if not callable(builder):
            return
        fields = builder(raw_probability, [("A", raw_probability)])
        self.assertEqual(fields["raw_probability"], raw_probability)
        self.assertEqual(
            fields["top3_candidates"][0]["raw_probability"],
            raw_probability,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path, json_path = save_prediction_logs(
                Path(temp_dir),
                [fields],
                {"raw_probability": raw_probability, "top3_candidates": fields["top3_candidates"]},
                stamp="exact",
            )
            with csv_path.open(encoding="utf-8-sig", newline="") as handle:
                csv_row = next(csv.DictReader(handle))
            json_payload = json.loads(json_path.read_text(encoding="utf-8"))

        self.assertEqual(float(csv_row["raw_probability"]), raw_probability)
        self.assertEqual(
            json.loads(csv_row["top3_candidates"])[0]["raw_probability"],
            raw_probability,
        )
        self.assertEqual(json_payload["raw_probability"], raw_probability)
        self.assertEqual(
            json_payload["top3_candidates"][0]["raw_probability"],
            raw_probability,
        )

    def test_one_record_team_export_is_readable_and_round_trips_probability(self):
        raw_probability = 0.12345678901234568
        expected_policy = {
            "kind": "uncalibrated_softmax",
            "acceptance_policy": "disabled_no_risk_coverage_evidence",
            "calibration_artifact": None,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            session = TeamTestSession(
                output_dir=Path(temp_dir),
                tester_id="tester",
                labels=["A"],
                label_display={"A": "甲"},
                trials_per_label=1,
                model_version="legacy",
            )
            session.confirm_prediction(
                predicted_label="A",
                raw_probability=raw_probability,
                top3_candidates=[("A", raw_probability)],
                clip_start_sec=1.25,
                clip_end_sec=2.5,
                finalize_sec=2.75,
            )

            paths = export_team_test_reports(session)
            with paths.trials_csv.open(encoding="utf-8-sig", newline="") as handle:
                csv_reader = csv.DictReader(handle)
                csv_row = next(csv_reader)
                csv_headers = csv_reader.fieldnames
            workbook = load_workbook(paths.workbook, read_only=True, data_only=True)
            sheet = workbook["逐次測試結果"]
            rows = list(sheet.iter_rows(values_only=True))
            header_cells = list(sheet[1])
            value_cells = list(sheet[2])
            xlsx_cells = {
                header.value: (cell.value, cell.data_type)
                for header, cell in zip(header_cells, value_cells)
            }
            workbook.close()

        self.assertIn("raw_probability", csv_headers)
        self.assertEqual(float(csv_row["raw_probability"]), raw_probability)
        self.assertEqual(json.loads(csv_row["probability_policy"]), expected_policy)
        self.assertEqual(
            json.loads(csv_row["top3_candidates"])[0]["raw_probability"],
            raw_probability,
        )
        xlsx_headers = list(rows[0])
        xlsx_row = dict(zip(xlsx_headers, rows[1]))
        self.assertEqual(float(xlsx_row["raw_probability"]), raw_probability)
        self.assertEqual(json.loads(xlsx_row["probability_policy"]), expected_policy)
        self.assertEqual(
            json.loads(xlsx_row["top3_candidates"])[0]["raw_probability"],
            raw_probability,
        )
        expected_numeric_cells = {
            "raw_probability": raw_probability,
            "clip_start_sec": 1.25,
            "clip_end_sec": 2.5,
            "finalize_sec": 2.75,
            "segment_duration_sec": 1.25,
            "finalize_delay_sec": 0.25,
        }
        for field, expected in expected_numeric_cells.items():
            with self.subTest(field=field):
                self.assertEqual(xlsx_cells[field], (expected, "n"))


class Knee42RawProbabilityTests(unittest.TestCase):
    def test_prediction_dataclass_names_raw_probability_with_deprecated_alias(self):
        field_names = {field.name for field in dataclasses.fields(knee42_ivcam.Prediction)}

        self.assertIn("raw_probability", field_names)
        self.assertNotIn("confidence", field_names)
        prediction = knee42_ivcam.Prediction(
            label_id="K42_01",
            display_text="你好",
            raw_probability=0.25,
        )
        with self.assertWarns(DeprecationWarning):
            self.assertEqual(prediction.confidence, prediction.raw_probability)

    def test_decode_logits_reports_exact_softmax_values_in_sorted_order(self):
        labels = [f"K42_{index:02d}" for index in range(1, 43)]
        logits = torch.tensor([[0.0, 3.0, 2.0] + [-1.0] * 39], dtype=torch.float32)
        expected = torch.softmax(logits[0], dim=0)

        result = knee42_ivcam.decode_logits(
            logits,
            labels,
            {label: label for label in labels},
        )

        expected_indices = torch.argsort(expected, descending=True)[:3].tolist()
        self.assertEqual(
            [item.label_id for item in result.top3],
            [labels[index] for index in expected_indices],
        )
        self.assertTrue(
            all(hasattr(item, "raw_probability") for item in result.top3),
            "decoded predictions must expose raw_probability",
        )
        for item, index in zip(result.top3, expected_indices):
            self.assertAlmostEqual(
                item.raw_probability,
                float(expected[index].item()),
                places=12,
            )
        self.assertAlmostEqual(
            sum(item.raw_probability for item in result.top3),
            float(expected[expected_indices].sum().item()),
            places=7,
        )
        self.assertAlmostEqual(float(expected.sum().item()), 1.0, places=6)

    def test_formal_overlay_and_structured_output_use_raw_probability_terms(self):
        prediction = knee42_ivcam.Prediction("K42_01", "你好", 0.037)
        result = knee42_ivcam.InferenceResult(
            top1=prediction,
            top3=(prediction, prediction, prediction),
        )

        rendered = "\n".join(
            knee42_ivcam.overlay_lines(
                result,
                fps=30.0,
                source="camera:0",
                mode="auto",
                state="RESULT",
            )
        )
        capture_source = inspect.getsource(knee42_ivcam.run_capture)

        self.assertIn("raw probability", rendered.lower())
        self.assertNotIn("confidence", rendered.lower())
        self.assertIn('"top1_raw_probability"', capture_source)
        self.assertNotIn('"top1_confidence"', capture_source)

    def test_formal_knee42_parser_has_no_acceptance_threshold_option(self):
        parser = knee42_ivcam.build_parser()
        option_strings = {
            option
            for action in parser._actions
            for option in action.option_strings
        }

        self.assertNotIn("--min-conf-override", option_strings)
        self.assertFalse(
            any("accept" in option or "confidence" in option for option in option_strings)
        )

    def test_decode_logits_has_no_acceptance_or_rejection_gate(self):
        source = inspect.getsource(knee42_ivcam.decode_logits).lower()

        for forbidden in ("threshold", "reject", "unknown", "accept"):
            with self.subTest(forbidden=forbidden):
                self.assertNotIn(forbidden, source)


if __name__ == "__main__":
    unittest.main()
