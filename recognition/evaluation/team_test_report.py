from __future__ import annotations

import csv
import json
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from recognition.evaluation.team_test_session import TeamTestSession
from recognition.realtime.probability_reporting import probability_policy_record


ERROR_REASON_OPTIONS = [
    "未啟動",
    "切太早",
    "切太晚／不結束",
    "切段正常但辨識錯",
    "自己做錯",
    "環境問題",
    "其他",
]


@dataclass(frozen=True)
class TeamReportPaths:
    workbook: Path
    trials_csv: Path
    session_json: Path


def build_label_summary(session: TeamTestSession) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for order, label_id in enumerate(session.labels):
        records = [record for record in session.records if record.expected_label == label_id]
        completed = len(records)
        correct = sum(record.top1_correct for record in records)
        incorrect = completed - correct
        confusion = Counter(
            record.predicted_label for record in records if not record.top1_correct
        )
        most_confused = confusion.most_common(1)[0][0] if confusion else ""
        rows.append(
            {
                "label_id": label_id,
                "sentence_text": session.label_display.get(label_id, label_id),
                "completed_trials": completed,
                "correct_trials": correct,
                "incorrect_trials": incorrect,
                "accuracy": (correct / completed) if completed else None,
                "most_confused_label": most_confused,
                "most_confused_text": session.label_display.get(
                    most_confused, most_confused
                ),
                "label_order": order,
            }
        )
    rows.sort(
        key=lambda row: (
            row["accuracy"] is None,
            row["accuracy"] if row["accuracy"] is not None else 2.0,
            row["label_order"],
        )
    )
    return rows


def _trial_row(record) -> dict[str, object]:
    row = asdict(record)
    row["top3_candidates"] = " | ".join(
        f"{item['label']}:{float(item['raw_probability']):.4f}"
        for item in record.top3_candidates
    )
    return row


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = list(rows[0].keys()) if rows else [
        "tester_id",
        "global_trial_number",
        "expected_label",
        "predicted_label",
        "top1_correct",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _append_rows(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = min(
            42,
            max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2),
        )
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _write_workbook(path: Path, session: TeamTestSession) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "不準句型總表"
    summary = build_label_summary(session)
    summary_headers = [
        "句型代號",
        "句型內容",
        "已完成次數",
        "正確次數",
        "錯誤次數",
        "正確率",
        "最常誤判代號",
        "最常誤判句型",
    ]
    _append_rows(
        summary_sheet,
        summary_headers,
        [
            [
                row["label_id"],
                row["sentence_text"],
                row["completed_trials"],
                row["correct_trials"],
                row["incorrect_trials"],
                row["accuracy"],
                row["most_confused_label"],
                row["most_confused_text"],
            ]
            for row in summary
        ],
    )
    for cell in summary_sheet["F"][1:]:
        cell.number_format = "0.0%"
    if summary_sheet.max_row >= 2:
        summary_sheet.conditional_formatting.add(
            f"F2:F{summary_sheet.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=0.8,
                mid_color="FFEB84",
                end_type="num",
                end_value=1,
                end_color="63BE7B",
            ),
        )

    trial_sheet = workbook.create_sheet("逐次測試結果")
    trial_rows = [_trial_row(record) for record in session.records]
    trial_headers = list(trial_rows[0].keys()) if trial_rows else [
        "tester_id",
        "global_trial_number",
        "expected_label",
        "predicted_label",
        "top1_correct",
    ]
    _append_rows(
        trial_sheet,
        trial_headers,
        [[row.get(header, "") for header in trial_headers] for row in trial_rows],
    )

    notes_sheet = workbook.create_sheet("組員備註")
    note_headers = [
        "全體次序",
        "預期代號",
        "預期句型",
        "預測代號",
        "預測句型",
        "錯誤原因",
        "補充備註",
    ]
    incorrect_records = [record for record in session.records if not record.top1_correct]
    _append_rows(
        notes_sheet,
        note_headers,
        [
            [
                record.global_trial_number,
                record.expected_label,
                record.expected_text,
                record.predicted_label,
                record.predicted_text,
                "",
                "",
            ]
            for record in incorrect_records
        ],
    )
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(ERROR_REASON_OPTIONS) + '"',
        allow_blank=True,
    )
    validation.error = "請從清單選擇錯誤原因"
    validation.errorTitle = "錯誤原因格式"
    notes_sheet.add_data_validation(validation)
    validation.add(f"F2:F{max(2, notes_sheet.max_row)}")

    workbook.save(path)


def export_team_test_reports(session: TeamTestSession) -> TeamReportPaths:
    session.session_dir.mkdir(parents=True, exist_ok=True)
    suffix = session.tester_id
    paths = TeamReportPaths(
        workbook=session.session_dir / f"team_results_{suffix}.xlsx",
        trials_csv=session.session_dir / f"team_trials_{suffix}.csv",
        session_json=session.session_dir / f"team_session_{suffix}.json",
    )
    trial_rows = [_trial_row(record) for record in session.records]
    _write_csv(paths.trials_csv, trial_rows)
    paths.session_json.write_text(
        json.dumps(
            {
                "schema_version": 2,
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "tester_id": session.tester_id,
                "model_version": session.model_version,
                "labels": session.labels,
                "trials_per_label": session.trials_per_label,
                "completed_trials": session.completed_trials,
                "total_trials": session.total_trials,
                "is_complete": session.is_complete,
                "runtime_metadata": session.runtime_metadata,
                "probability_policy": probability_policy_record(),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    _write_workbook(paths.workbook, session)
    return paths


def package_team_results(session_dir: Path, output_zip: Path) -> Path:
    session_dir = Path(session_dir).resolve()
    output_zip = Path(output_zip).resolve()
    allowed_patterns = ["team_results_*.xlsx", "team_trials_*.csv", "team_session_*.json"]
    report_files: list[Path] = []
    for pattern in allowed_patterns:
        matches = sorted(path for path in session_dir.glob(pattern) if path.is_file())
        if len(matches) != 1:
            raise FileNotFoundError(
                f"expected exactly one report matching {pattern}, found {len(matches)}"
            )
        report_files.extend(matches)
    output_zip.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for report_file in report_files:
            archive.write(report_file, arcname=report_file.name)
    return output_zip
